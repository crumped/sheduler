from django.shortcuts import render
import openpyxl
import random
import datetime
import calendar
from math import ceil
import copy


# TODO
# dokładniej sprawdzać dni


def index(request):
    if "GET" == request.method:
        return render(request, 'api/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file, data_only=True)

        worksheet = wb["Arkusz1"]
        worksheet2 = wb["Arkusz2"]
        worksheet3 = wb["Arkusz3"]

        excel_data = list()
        excel_data_holiday = {}
        excel_data_cannot_work = {}

        for index, row in enumerate(worksheet2.iter_rows(), start=0):
            excel_data_holiday[str(row[0].value)] = []
            for index2, cell in enumerate(row, start=0):
                if index2 != 0:
                    excel_data_holiday[str(row[0].value)].append(cell.value)

        for index, row in enumerate(worksheet3.iter_rows(), start=0):
            excel_data_cannot_work[str(row[0].value)] = []
            for index2, cell in enumerate(row, start=0):
                if index2 != 0:
                    excel_data_cannot_work[str(row[0].value)].append(cell.value)

        # print(excel_data_holiday)
        today = datetime.datetime.today()

        month, year = next_mouth(today.month, today.year)
        num_days = calendar.monthrange(year, month)[1]
        days = [{"day": datetime.date(year, month, day), "day_name": day, "users_morning": [], "users_day": [],
                 "users_night": [], "users_afternoon": [], "holidays": [], "cannot_work": []} for day in range(1, num_days + 1)]
        create_data(excel_data, worksheet, days, excel_data_holiday)
        while True:
            print("==========================================================================================================")
            print("==========================================   Reset   =====================================================")
            print("==========================================================================================================")
            users = copy.deepcopy(excel_data)
            days_tmp = copy.deepcopy(days)
            get_schedule(days_tmp, users, excel_data_holiday, excel_data_cannot_work)
            make_up_all_days(days_tmp, users, excel_data_holiday, excel_data_cannot_work)
            make_up_all_mornings(days_tmp, users, excel_data_holiday, excel_data_cannot_work)
            print("days_tmp")
            print(days_tmp)
            print("users")
            print(users)
            if check_data(users):
                excel_data = users
                days = days_tmp
                break

        representant = representant_data(days, users)

        # print(days)
        # print(excel_data)
        # print(users)

        return render(request, 'api/index.html', {"excel_data": excel_data, "days": days, "representant": representant})


def representant_data(days, users):
    representant = []
    users_data = users.copy()
    users_data.pop(0)
    for i, user in enumerate(users_data, start=0):
        day_list = [user['user_id'], user['username']]
        for j, day in enumerate(days, start=0):
            if int(user['user_id']) in day['users_day']:
                day_list.append("D")
            elif int(user['user_id']) in day['users_night']:
                day_list.append("N")
            elif int(user['user_id']) in day['users_afternoon']:
                day_list.append("Płd")
            elif int(user['user_id']) in day['users_morning']:
                day_list.append("Ra")
            elif int(user['user_id']) in day['holidays']:
                day_list.append("U")
            elif int(user['user_id']) in day['cannot_work']:
                day_list.append("X")
            else:
                day_list.append(" ")

        representant.append(day_list)
    return representant


def check_data(users):
    users_data = users.copy()
    users_data.pop(0)
    for user in users_data:
        if user['month'] != user['total_hours'] or user['12'] != int(user['total_hours'] / 12):
            return False
        decimal = user['total_hours'] / 12
        number = round(decimal, 2)
        number_dec = float(str(number - int(number))[1:])
        if round(number_dec, 2) == 0.67:
            if user['8'] != 1:
                return False
        if round(number_dec, 2) == 0.33:
            if user['4'] != 1:
                return False
    return True


def check_8or4(users):
    users_data = users.copy()
    users_data.pop(0)
    for user in users_data:
        decimal = user['total_hours'] / 12
        number = round(decimal, 2)
        number_dec = float(str(number - int(number))[1:])
        if round(number_dec, 2) == 0.67:
            if user['8'] != 1:
                return False
        if round(number_dec, 2) == 0.33:
            if user['4'] != 1:
                return False
    return True


def check_12(users):
    users_data = users.copy()
    users_data.pop(0)
    for user in users_data:
        if user['12'] != int(user['total_hours'] / 12):
            return False
    return True


def create_data(excel_data, worksheet, days, holidays):
    for index, row in enumerate(worksheet.iter_rows(), start=0):
        row_data = list()
        if index == 0:
            row_data.append('Lp')
            for cell in row:
                row_data.append(str(cell.value))
        else:
            weeks = {}
            if row[0].value in holidays:
                total_time = (int(row[1].value) - (8 * len(holidays[row[0].value])))
            else:
                total_time = row[1].value
            row_data = {"user_id": str(index), "username": row[0].value, "total_hours": total_time, "month": 0,
                        "12": 0, "8": 0, "4": 0}
            for day in days:
                week = week_of_month(day['day'])
                if week not in weeks:
                    weeks[week] = 0
            row_data['weeks'] = weeks
        excel_data.append(row_data)


def week_of_month(dt):
    """ Returns the week of the month for the specified date.
    """

    first_day = dt.replace(day=1)

    dom = dt.day
    adjusted_dom = dom + first_day.weekday()

    return int(ceil(adjusted_dom / 7.0))


def next_mouth(month, year):
    if month == 12:
        year = year + 1
        month = 1
    else:
        month = month + 1

    return month, year


def get_nights(day, i, days, users, available_users):
    users_last_night = []
    users_id_hours = check_user_hours(available_users, day, 12)
    if i != 0:
        users_last_night = days[i - 1]['users_night']
    users_id = list(set(users_last_night) | set(users_id_hours))
    delete_user(available_users, users_id)
    users_id = []
    if len(available_users) > 2:
        random_numbers = random.sample(range(0, len(available_users)), 3)
        week = week_of_month(day['day'])
        for number in random_numbers:
            user_id = int(available_users[number]["user_id"])
            users_id.append(user_id)
            users[user_id]['weeks'][week] = users[user_id]['weeks'][week] + 12
            users[user_id]['month'] = users[user_id]['month'] + 12
            users[user_id]['12'] = users[user_id]['12'] + 1
    days[i]['users_night'] = users_id


def get_days(day, i, days, users, available_users):
    users_id_hours = check_user_hours(available_users, day, 12)
    if i != 0:
        users_id = day['users_night'] + days[i - 1]['users_night'] + days[i - 1]['users_afternoon']
        users_b = []
        users_c = []
        users_f = []
        if i < len(days) - 2:
            users_c = get_workers(days[i - 1], days[i + 1])
        if i < len(days) - 3:
            users_f = get_workers(days[i + 1], days[i + 2])
        if i > 1:
            users_b = get_workers(days[i - 1], days[i - 2])
        users_id = list(set(users_id) | set(users_b) | set(users_c) | set(users_f))
    else:
        users_id = day['users_night']
        users_b = get_workers(days[i + 1], days[i + 2])
        users_id = list(set(users_id) | set(users_b))

    users_id = list(set(users_id) | set(users_id_hours))
    delete_user(available_users, users_id)
    users_id = []
    if len(available_users) > 2:
        random_numbers = random.sample(range(0, len(available_users)), 3)
        week = week_of_month(day['day'])
        for number in random_numbers:
            user_id = int(available_users[number]["user_id"])
            users_id.append(user_id)
            users[user_id]['weeks'][week] = users[user_id]['weeks'][week] + 12
            users[user_id]['month'] = users[user_id]['month'] + 12
            users[user_id]['12'] = users[user_id]['12'] + 1
    days[i]["users_day"] = users_id


def get_afternoons(day, i, days, users, available_users):
    users_afternoons = []
    if day['day'].weekday() == 1 or day['day'].weekday() == 4:
        users_id_hours = check_user_8_4_hours(available_users, day)
        if i != 0:
            users_id = day['users_night'] + days[i - 1]['users_night'] + day['users_day'] + days[i + 1]['users_day']
        else:
            users_id = day['users_night'] + day['users_day'] + days[i + 1]['users_day']
        users_id = list(set(users_id) | set(users_id_hours))
        delete_user(available_users, users_id)
        users_id = []
        if len(available_users) > 2:
            random_numbers = random.sample(range(0, len(available_users)), 1)
            week = week_of_month(day['day'])
            for number in random_numbers:
                user_id = int(available_users[number]["user_id"])
                users_id.append(user_id)

                decimal = users[user_id]['total_hours'] / 12
                number = round(decimal, 2)
                number_dec = float(str(number - int(number))[1:])
                if round(number_dec, 2) == 0.67:
                    users[user_id]['weeks'][week] = users[user_id]['weeks'][week] + 8
                    users[user_id]['month'] = users[user_id]['month'] + 8
                    users[user_id]['8'] = users[user_id]['8'] + 1
                if round(number_dec, 2) == 0.33:
                    users[user_id]['weeks'][week] = users[user_id]['weeks'][week] + 4
                    users[user_id]['month'] = users[user_id]['month'] + 4
                    users[user_id]['4'] = users[user_id]['4'] + 1
                users_afternoons.append(user_id)
        days[i]["users_afternoon"] = users_id
    else:
        days[i]["users_afternoon"] = []


def delete_user(users, indexes):
    array_indexes = []
    for i in range(len(users)):
        if int(users[i]['user_id']) in indexes:
            array_indexes.append(i)

    for index in sorted(array_indexes, reverse=True):
        del users[index]


def check_user_hours(users, day, number, is_not_enought=False):
    users_id = []
    for i, user in enumerate(users, start=0):
        week = week_of_month(day['day'])
        if is_not_enought:
            max_lim = 48
        else:
            max_lim = 42
        if user['weeks'][week] + number > max_lim:
            users_id.append(int(user['user_id']))

        if user['month'] + number > user['total_hours']:
            if int(user['user_id']) not in users_id:
                users_id.append(int(user['user_id']))

    return users_id


def check_user_8_4_hours(users, day, is_not_enought=False):
    users_id = []
    for i, user in enumerate(users, start=0):
        decimal = user['total_hours'] / 12
        number = round(decimal, 2)
        number_dec = float(str(number - int(number))[1:])
        if number_dec == 0.0:
            users_id.append(int(user['user_id']))

        if round(number_dec, 2) == 0.67:
            if user['8'] == 1:
                users_id.append(int(user['user_id']))
            week = week_of_month(day['day'])
            if is_not_enought:
                max_lim = 48
            else:
                max_lim = 42
            if user['weeks'][week] + 8 > max_lim:
                users_id.append(int(user['user_id']))
            if user['month'] + 8 > user['total_hours']:
                if int(user['user_id']) not in users_id:
                    users_id.append(int(user['user_id']))

        if round(number_dec, 2) == 0.33:
            if user['4'] == 1:
                users_id.append(int(user['user_id']))
            week = week_of_month(day['day'])
            if is_not_enought:
                max_lim = 48
            else:
                max_lim = 42
            if user['weeks'][week] + 4 > max_lim:
                users_id.append(int(user['user_id']))
            if user['month'] + 4 > user['total_hours']:
                if int(user['user_id']) not in users_id:
                    users_id.append(int(user['user_id']))

    return users_id


def get_schedule(days, users, holidays, cannot_work):
    for i, day in enumerate(days, start=0):
        users_data = users.copy()
        users_data.pop(0)
        available_users = users_data.copy()
        #TODO delete users from available users list for that day
        user_id = []
        for key in holidays:
            if day['day_name'] in holidays[key]:
                for user in available_users:
                    if user['username'] == key:
                        user_id.append(int(user['user_id']))
                        if int(user['user_id']) not in days[i]['holidays']:
                            days[i]['holidays'].append(int(user['user_id']))
        for key in cannot_work:
            if day['day_name'] in cannot_work[key]:
                for user in available_users:
                    if user['username'] == key:
                        user_id.append(int(user['user_id']))
                        if int(user['user_id']) not in days[i]['cannot_work']:
                            days[i]['cannot_work'].append(int(user['user_id']))

        delete_user(available_users, user_id)
        get_nights(day, i, days, users, available_users)
        get_days(day, i, days, users, available_users)
        get_afternoons(day, i, days, users, available_users)
        print("days")
        print(days)


def make_up_all_days(days, users, holidays, cannot_work):
    rols = 0
    while not check_12(users):
        rols = rols + 1
        if rols == 100:
            break
        for i, day in enumerate(days, start=0):
            if day['day'].weekday() != 5 or day['day'].weekday() != 6:
                users_id = []
                users_data = users.copy()
                users_data.pop(0)
                available_users = users_data.copy()
                # TODO delete users from available users list for that day
                holidays_user_id = []
                for key in holidays:
                    if day['day_name'] in holidays[key]:
                        for user in available_users:
                            if user['username'] == key:
                                holidays_user_id.append(int(user['user_id']))
                                if int(user['user_id']) not in days[i]['holidays']:
                                    days[i]['holidays'].append(int(user['user_id']))

                for key in cannot_work:
                    if day['day_name'] in cannot_work[key]:
                        for user in available_users:
                            if user['username'] == key:
                                holidays_user_id.append(int(user['user_id']))
                                if int(user['user_id']) not in days[i]['cannot_work']:
                                    days[i]['cannot_work'].append(int(user['user_id']))

                users_delete_id = find_completed_users(available_users)
                if rols > 50:
                    users_id_hours = check_user_hours(available_users, day, 12)
                else:
                    users_id_hours = check_user_hours(available_users, day, 12, True)
                if i == 0:
                    users_id = day['users_night'] + day['users_day'] + day['users_afternoon']
                    users_b = get_workers(days[i + 1], days[i + 2])
                    users_id = list(set(users_id) | set(users_b))
                else:
                    users_id = day['users_night'] + days[i - 1]['users_night'] + day['users_day']\
                               + day['users_afternoon'] + days[i - 1]['users_afternoon']
                    users_b = []
                    users_c = []
                    users_f = []
                    if i < len(days) - 2:
                        users_c = get_workers(days[i - 1], days[i + 1])
                    if i < len(days) - 3:
                        users_f = get_workers(days[i + 1], days[i + 2])
                    if i > 1:
                        users_b = get_workers(days[i - 1], days[i - 2])
                    users_id = list(set(users_id) | set(users_b) | set(users_c) | set(users_f))

                users_id = list(set(users_id) | set(users_delete_id) | set(users_id_hours) | set(holidays_user_id))
                delete_user(available_users, users_id)
                if len(available_users) > 0:
                    number = random.randint(0, len(available_users) - 1)
                    week = week_of_month(day['day'])
                    user_id = int(available_users[number]["user_id"])
                    users[user_id]['weeks'][week] = users[user_id]['weeks'][week] + 12
                    users[user_id]['month'] = users[user_id]['month'] + 12
                    users[user_id]['12'] = users[user_id]['12'] + 1
                    days[i]["users_day"].append(user_id)


def make_up_all_mornings(days, users, holidays, cannot_work):
    rols = 0
    while not check_8or4(users):
        rols = rols + 1
        if rols == 100:
            break
        for i, day in enumerate(days, start=0):
            if day['day'].weekday() == 0 or day['day'].weekday() == 2 or day['day'].weekday() == 3:
                users_id = []
                users_data = users.copy()
                users_data.pop(0)
                available_users = users_data.copy()
                # TODO delete users from available users list for that day
                holidays_user_id = []
                for key in holidays:
                    if day['day_name'] in holidays[key]:
                        for user in available_users:
                            if user['username'] == key:
                                holidays_user_id.append(int(user['user_id']))
                                if int(user['user_id']) not in days[i]['holidays']:
                                    days[i]['holidays'].append(int(user['user_id']))

                for key in cannot_work:
                    if day['day_name'] in cannot_work[key]:
                        for user in available_users:
                            if user['username'] == key:
                                holidays_user_id.append(int(user['user_id']))
                                if int(user['user_id']) not in days[i]['cannot_work']:
                                    days[i]['cannot_work'].append(int(user['user_id']))

                # delete_user(available_users, holidays_user_id)
                if rols > 50:
                    users_id_hours = check_user_8_4_hours(available_users, day)
                else:
                    users_id_hours = check_user_8_4_hours(available_users, day, True)

                if i == 0:
                    users_id = day['users_night'] + day['users_day'] + day['users_afternoon']
                    users_b = get_workers(days[i + 1], days[i + 2])
                    users_id = list(set(users_id) | set(users_b))
                else:
                    users_id = day['users_night'] + days[i - 1]['users_night'] + day['users_day'] + day['users_afternoon'] + days[i - 1]['users_afternoon']
                    users_b = []
                    users_c = []
                    users_f = []
                    if i < len(days) - 2:
                        users_c = get_workers(days[i - 1], days[i + 1])
                    if i < len(days) - 3:
                        users_f = get_workers(days[i + 1], days[i + 2])
                    if i > 1:
                        users_b = get_workers(days[i - 1], days[i - 2])
                    users_id = list(set(users_id) | set(users_b) | set(users_c) | set(users_f))

                users_id = list(set(users_id) | set(users_id_hours) | set(holidays_user_id))
                delete_user(available_users, users_id)
                print("available_users")
                print(available_users)
                if len(available_users) > 0:
                    number = random.randint(0, len(available_users) - 1)
                    week = week_of_month(day['day'])
                    user_id = int(available_users[number]["user_id"])
                    decimal = users[user_id]['total_hours'] / 12
                    number = round(decimal, 2)
                    number_dec = float(str(number - int(number))[1:])
                    if round(number_dec, 2) == 0.67:
                        users[user_id]['weeks'][week] = users[user_id]['weeks'][week] + 8
                        users[user_id]['month'] = users[user_id]['month'] + 8
                        users[user_id]['8'] = users[user_id]['8'] + 1
                    if round(number_dec, 2) == 0.33:
                        users[user_id]['weeks'][week] = users[user_id]['weeks'][week] + 4
                        users[user_id]['month'] = users[user_id]['month'] + 4
                        users[user_id]['4'] = users[user_id]['4'] + 1
                    days[i]["users_morning"] = [user_id]


def get_workers(day_1, day_2):
    return set(day_1['users_day']) - (set(day_1['users_day']) - set(day_2['users_day']))
    # return set(part) - (set(part) - set(day_2['users_afternoon']))


def find_completed_users(users):
    users_id = []
    for user in users:
        if user['12'] == int(user['total_hours'] / 12):
            users_id.append(int(user['user_id']))
    return users_id
